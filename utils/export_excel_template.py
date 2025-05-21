from openpyxl import load_workbook
from openpyxl.drawing.image import Image
import os
import win32com.client as win32

def isi_template_excel(df_pred, template_path="static/template_download.xlsx", output_path="static/hasil/data_prediksi.xlsx"):
    os.makedirs("static/hasil", exist_ok=True)
    wb = load_workbook(template_path)
    ws = wb.active

    # Isi data mulai dari baris 7
    start_row = 7
    for i, row in enumerate(df_pred.itertuples(index=False), start=start_row):
        ws[f"A{i}"] = row.hari
        ws[f"B{i}"] = row.tanggal
        ws[f"C{i}"] = round(row.RH_avg, 2)
        ws[f"D{i}"] = round(row.Tavg, 2)
        ws[f"E{i}"] = round(row.RR, 2)
        ws[f"F{i}"] = round(row.ss, 2)
        ws[f"G{i}"] = row.klasifikasi
        ws[f"H{i}"] = "\n".join(row.keterangan) if isinstance(row.keterangan, list) else row.keterangan

    wb.save(output_path)
    print("✅ Laporan dengan logo berhasil disimpan:", output_path)